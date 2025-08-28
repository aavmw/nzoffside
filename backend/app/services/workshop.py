import re
import json 
from psycopg_pool import ConnectionPool
import io
from datetime import datetime
from google.oauth2 import service_account
from googleapiclient.discovery import build, HttpError
from googleapiclient.http import MediaIoBaseDownload
from zoneinfo import ZoneInfo
from openpyxl import load_workbook

from ..config import get_settings

settings = get_settings()
MASTER_SPREADSHEET_ID = settings.MASTER_SPREADSHEET_ID

class DbManager:
    def __init__(self, row_factory=None):
        # SQLAlchemy uses "postgresql+psycopg://"; raw psycopg uses "postgresql://"
        url = settings.DATABASE_URL
        if url.startswith("postgresql+psycopg://"):
            url = url.replace("postgresql+psycopg://", "postgresql://", 1)

        self.url = url
        self.row_factory = row_factory
        self.db_pool: ConnectionPool | None = None

    def __enter__(self):
        # min_size/max_size mirror your old SimpleConnectionPool(1, 5)
        self.db_pool = ConnectionPool(
            conninfo=self.url,
            min_size=1,
            max_size=5,
            timeout=30,          # seconds to wait for a free conn
            kwargs={},           # extra psycopg.connect kwargs if needed
        )
        return self

    def __exit__(self, exc_type, exc, tb):
        if self.db_pool:
            self.db_pool.close()

    def execute_query(self, query: str, params: tuple | None = None, fetch_results: bool = False):
        if not self.db_pool:
            raise RuntimeError("DbManager used without context manager")

        with self.db_pool.connection() as conn:
            cursor_kwargs = {}
            if self.row_factory:
                cursor_kwargs["row_factory"] = self.row_factory

            with conn.cursor(**cursor_kwargs) as cur:
                cur.execute(query, params or ())
                if fetch_results:
                    return cur.fetchall(), cur.description
                conn.commit()
                return cur.rowcount


class GoogleApiService:

    SCOPES = ['https://www.googleapis.com/auth/drive',
              'https://www.googleapis.com/auth/spreadsheets'
              ]
    
    with open(settings.GOOGLE_CREDS_PATH) as f:
        CREDS = json.loads(f.read())

    def __init__(self):
        self.credentials = service_account.Credentials.from_service_account_info(self.CREDS, scopes=self.SCOPES)

    @property
    def drive_service(self):
        return build('drive', 'v3', credentials=self.credentials)

    @property
    def sheets_service(self):
        return build('sheets', 'v4', credentials=self.credentials)
    
    def get_folder_files_info(self, query: str) -> list:

        page_token = None
        result = []

        while True:

            response = self.drive_service.files().list(
                    q=query,
                    fields="nextPageToken, files(id, name, modifiedTime)",
                    pageToken=page_token,
                    pageSize=100
                ).execute()

            result.extend(response.get('files', []))
            page_token = response.get('nextPageToken', None)

            if not page_token:
                break

        return result
     
class OperationManager:

    COLORS_DICT = {
            'completed' : {"red": 0.7, "green": 0.9, "blue": 0.7},
            'in_work' : {"red": 1.0, "green": 0.9, "blue": 0.6},
            'pending' : {"red": 1.0, "green": 0.65, "blue": 0.1}
        }

    def __init__(self):
        pass

    def _save_log(self, message: dict):
        """Save operation log to database"""

        query = '''
                INSERT INTO operation_log (
                    message_dttm,
                    message
                )
                VALUES (%s, %s)
        '''
        
        params = (
            datetime.now(),
            json.dumps(message)
        )
        
        with DbManager() as db:
            db.execute_query(query, params)

    def _update_operation(self, message: dict):
        """Update operation data in JSON column"""
        drive_id = message.get('jobCardCode')
        operation = message.get('operation')

        if not drive_id or not operation:
        # Nothing sensible to update
            return
        
        operation_data = {}
        mapping = {
            'startDateTime': 'start_dttm',
            'endDateTime': 'end_dttm',
            'email': 'user',
            'comment': 'comment',
            'used_mnhrs' : 'used_mnhrs'
        }
        
        for src, dst in mapping.items():
            if src in message:
                operation_data[dst] = message.get(src)

        if not operation_data:
            return       

        query = '''
                UPDATE job_cards
                SET operations = jsonb_set(
                    operations::jsonb,                     -- cast to jsonb
                    ARRAY[%s],
                    COALESCE((operations::jsonb) -> %s, '{}'::jsonb) || %s::jsonb,
                    true
                )::json                                    -- cast back if the column type is json
                WHERE drive_id = %s;
            '''
        
        params = (
            operation,               # path element
            operation,               # lookup existing node
            json.dumps(operation_data),      # only the keys you’re updating (e.g., {"comment":"test"})
            drive_id
        )
                
        with DbManager() as db:
            db.execute_query(query, params)

    def update_operation(self, message:dict):

        self._update_operation(message)
        self._save_log(message)

    def _get_operations_data(self):

        query = 'SELECT * FROM job_cards WHERE is_active = True ORDER BY project, name'

        with DbManager() as db:
            job_cards_table, description = db.execute_query(query, fetch_results=True)
            columns = [col[0] for col in description] 
        
        return [dict(zip(columns, row)) for row in job_cards_table]

    def create_sheets_dataset(self) -> list[list[str]]:

        db_data = self._get_operations_data()

        color_dict = self.COLORS_DICT
        

        result = [
            ['project', 'creation_date', 'name', 'part_number', 'serial_number', 'operations']
        ]

        colors_column_offset = len(result[0]) - 1

        current_dttm_columns = ['', 'last_update', datetime.now(tz=ZoneInfo("Europe/Moscow")).strftime('%d.%m.%Y %H:%M')]
        result[0].extend(current_dttm_columns)

        hyperlinks = {}
        colors = {}
        comments = {}

        for row_index, row in enumerate(db_data, start=1):

            operations = row['operations']

            result.append([
                row['project'],
                row['creation_dttm'].strftime('%Y-%m-%d %H:%M'),
                '',
                row['part_number'],
                row['serial_number'],
                *sorted(operations)
            ])

            hyperlinks[(row_index, 2)] = (row['name'], f"https://docs.google.com/spreadsheets/d/{row['drive_id']}")
            for col_index, op in enumerate(sorted(operations.keys())):
                if operations[op]['end_dttm']:
                    colors[(row_index, col_index + colors_column_offset)] = color_dict['completed']
                    if op != 'PPK':
                        colors[(row_index, col_index + colors_column_offset + 1)] = color_dict['pending']
                    else:
                        for i in range(6):
                            colors[(row_index, i)] = color_dict['completed']
                elif operations[op]['start_dttm']:
                    colors[(row_index, col_index + colors_column_offset)] = color_dict['in_work']

                if operations[op].get('comment'):
                    comments[(row_index, col_index + colors_column_offset)] = operations[op]['comment']

        return result, hyperlinks, colors, comments
    
    def get_job_card_data(self, drive_id):

        query = 'SELECT * FROM job_cards WHERE drive_id = %s'

        with DbManager() as db:
            job_card, description = db.execute_query(query, (drive_id, ), True)
            columns = [col[0] for col in description] 
            return dict(zip(columns, map(lambda x: str(x) if isinstance(x, datetime) else x, job_card[0])))


    def get_single_operation_data(self, drive_id, operation):

        db_operations = self.get_job_card_data(drive_id)['operations']

        if db_operations:
            single_operation_data = db_operations.get(operation)

            if single_operation_data:
                return single_operation_data

class JobCard:

    VALUES_MAP = {
        'name' : '7_1',
        'part_number' : '9_1',
        'serial_number' : '9_7',
        'project' : '5_7'
    }
    
    def __init__(self, drive_id: str, project: str, modified_time: datetime =None ):
        self.drive_id = drive_id
        self.project = project
        self._modified_time = modified_time
        self._data_cache = None
        self._extracted_operations = None
        self._google_api = GoogleApiService()
        self._sheets_service = self._google_api.sheets_service
        self._drive_service = self._google_api.drive_service

    def get_mapping(self, key):
        return self.VALUES_MAP.get(key)
    
    @property
    def modified_time(self):
        if self._modified_time is None:
            file_metadata = self._drive_service.files().get(
                fileId=self.drive_id,
                fields='modifiedTime'
            ).execute()
            self._modified_time = datetime.strptime(file_metadata['modifiedTime'], '%Y-%m-%dT%H:%M:%S.%fZ')
        return self._modified_time
          
    @property
    def data(self) -> dict:
        '''
        Download and parse GoogleSheet job card file.
        Returns dict with rowIndex_colIndex as key and value.
        '''
        if self._data_cache is None:

            try:
                values = self._sheets_service.spreadsheets().values().get(
                    spreadsheetId=self.drive_id,
                    range='JC'
                ).execute().get('values', [])
            except HttpError as error:
                if error.status_code == 400:
                    values = self._convert_xlsx_to_sheets()

            result = {}
            for row_idx, row in enumerate(values, start=1):
                for col_idx, value in enumerate(row, start=1):
                    result[f'{row_idx}_{col_idx}'] = value if value else None
            
            self._data_cache = result  
        return self._data_cache
    
    def _convert_xlsx_to_sheets(self):

        # Download file bytes
        request = self._drive_service.files().get_media(fileId=self.drive_id)
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)

        done = False
        while not done:
            status, done = downloader.next_chunk()

        fh.seek(0)

        # Load workbook from bytes
        wb = load_workbook(filename=fh, data_only=True)
        ws = wb.active  # first sheet

        # Extract rows as list of lists
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(list(row))

        return data
        
    def update_job_card_info(self):

        params = (
            self.drive_id,
            self.data.get(self.get_mapping('name')),
            str(self.modified_time),
            self.data.get(self.get_mapping('part_number')),
            self.data.get(self.get_mapping('serial_number')),
            str(self.modified_time),
            self.operations_produce(),
            self.project,
            True,
        )

        placeholders = ','.join(['%s'] * len(params))

        update_query = f'''
                        INSERT INTO job_cards (
                            drive_id, 
                            name, 
                            creation_dttm, 
                            part_number, 
                            serial_number, 
                            modified_dttm, 
                            operations,
                            project, 
                            is_active) 
                        VALUES ({placeholders})
                        ON CONFLICT(drive_id) DO UPDATE SET 
                            name = excluded.name,
                            part_number = excluded.part_number,
                            serial_number = excluded.serial_number,
                            modified_dttm = excluded.modified_dttm,
                            operations = excluded.operations,
                            project = excluded.project,
                            is_active = excluded.is_active;
                    '''

        with DbManager() as db:
            db.execute_query(update_query, params)
    
    @property    
    def extracted_operations(self) -> str:

        if self._extracted_operations is None:

            operations = {}
            operation_fields = {
                'start_dttm' : None,
                'end_dttm' : None,
                'comment' : None,
                'user' : None,
                'used_mnhrs': None
            }
            collection = False

            for idx in filter(lambda x: x.endswith('_1'), self.data):

                value = self.data[idx]

                if not value:
                    continue

                if re.search('End date', value):
                    break

                if collection:
                    value = str(value).replace('\n', '')
                    operations[value] = operation_fields
                
                if re.search('Start date', value):
                    collection = True

            operations['PPK'] = operation_fields

            self._extracted_operations = json.dumps(operations)

        return self._extracted_operations
    
    def operations_produce(self):
        '''
        Create operations JSON. 
        If JobCard already in DataBase make compare and
        keep DataBase stored values.
        '''

        db_operations_query = 'SELECT operations FROM job_cards WHERE drive_id = %s'
        
        with DbManager() as db:
            db_operations, _ = db.execute_query(db_operations_query, (self.drive_id,), True)
            try:
                db_operations = db_operations[0][0]
            except IndexError:
                db_operations = None

        drive_operations = json.loads(self.extracted_operations)


        if not db_operations or db_operations == drive_operations:
            return self.extracted_operations
        
        result = {}

        for operation in drive_operations:
            
            db_value = db_operations.get(operation)
            extracted_value = drive_operations.get(operation)

            if db_value:
                result[operation] = db_value
            else:
                result[operation] = extracted_value

        return json.dumps(result)

def update_db_job_cards_info():
    '''
    Function for check difference in Google drive folder modified time
    and DataBase modified time. If difference exists, triggers update
    job card info in DataBase.
    '''

    PROJECTS_FOLDER = '1pu8ssI1HI_5k8ewhufrO8HcoLDon2jVR'
    _google_api = GoogleApiService()

    in_work_folders = []
    job_cards_list = []
    current_db_data_query = 'SELECT drive_id, modified_dttm, is_active FROM job_cards'

    projects_folders = _google_api.get_folder_files_info(f"'{PROJECTS_FOLDER}' in parents and trashed = false")

    for parent in projects_folders:
        project = parent['name']
        temp_folders = _google_api.get_folder_files_info(f"'{parent['id']}' in parents and mimeType='application/vnd.google-apps.folder' and name contains 'in_work'")
        in_work_folders.extend(list(map(lambda x: x | {'project' : project}, temp_folders)))

    for folder in in_work_folders:
        project = folder['project']
        temp_folders = _google_api.get_folder_files_info(f"'{folder['id']}' in parents and trashed = false")
        job_cards_list.extend(list(map(lambda x: x | {'project' : project}, temp_folders)))

    with DbManager() as db:
        db_response, columns = db.execute_query(current_db_data_query, fetch_results=True)
        db_dict = {row[0] : {'modifiedTime' : row[1], 'is_active' : row[2]} for row in db_response}

    to_update_job_cards = []

    for file in job_cards_list:

        file['modifiedTime'] = datetime.strptime(file['modifiedTime'], '%Y-%m-%dT%H:%M:%S.%fZ')

        try:
            db_instance = db_dict.get(file['id'])
            db_instance_modified_time, db_instance_is_active = db_instance.get('modifiedTime'), db_instance.get('is_active')
        except AttributeError:
            to_update_job_cards.append(file)
            continue

        if db_instance_modified_time != file['modifiedTime'] or not db_instance_is_active:
            to_update_job_cards.append(file)

    #print(f'to update: {to_update_job_cards}')

    for file in to_update_job_cards:
        jc = JobCard(file['id'], file['project'], file['modifiedTime'])
        try:
            jc.update_job_card_info()
        except HttpError:
            print(f'failed: {file}')

    
    drive_ids = tuple(map(lambda x: x['id'], job_cards_list))
    to_set_inactive = [file_id for file_id in db_dict if file_id not in drive_ids]

    if to_set_inactive:
        placeholders = ','.join(['%s'] * len(to_set_inactive))
        inactive_query = f'''
            UPDATE job_cards
            SET is_active = False
            WHERE drive_id IN ({placeholders})
        '''
        with DbManager() as db:
            db.execute_query(inactive_query, tuple(to_set_inactive))

    #print(f'to inactive: {to_set_inactive}')

def _get_master_table_sheet_id():

    sheets_service = GoogleApiService().sheets_service

    sheet_metadata = sheets_service.spreadsheets().get(
        spreadsheetId=MASTER_SPREADSHEET_ID,
        fields="sheets(properties(sheetId,title))"
    ).execute()

    return next(
        sheet['properties']['sheetId'] 
        for sheet in sheet_metadata['sheets']
        if sheet['properties']['title'] == 'master'
    )


def update_google_master_table():

    sheets_service = GoogleApiService().sheets_service
    sheets_data, hyperlink_map, colors, comments = OperationManager().create_sheets_dataset()

    sheet_id = _get_master_table_sheet_id()

    values = []
    for row in sheets_data:
        sheets_row = []
        for cell in row:
            sheets_row.append({"userEnteredValue": {"stringValue": str(cell)}} if cell is not None else {})
        values.append(sheets_row)

    requests = [
        {
            "updateCells": {
                "range": {"sheetId": sheet_id},
                "fields": "userEnteredValue,userEnteredFormat.backgroundColor,userEnteredFormat.textFormat.foregroundColor,note"
                }
        },
        # 1. Update all cell values
        {
            "updateCells": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": len(sheets_data),
                    "startColumnIndex": 0,
                    "endColumnIndex": max([len(row) for row in sheets_data]) if sheets_data else 0
                },
                "rows": [
                        {
                            "values": [
                                {"userEnteredValue": {"stringValue": str(cell)}}
                                for cell in row
                            ]
                        }
                        for row in sheets_data
                    ],
                "fields": "userEnteredValue"
            }
        },
        # 2. Apply operation colors
        *[{
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_idx,
                    "endRowIndex": row_idx + 1,
                    "startColumnIndex": col_idx,
                    "endColumnIndex": col_idx + 1
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": color
                    }
                },
                "fields": "userEnteredFormat.backgroundColor"
            }
        } for (row_idx, col_idx), color in colors.items()],
        # 3. Add name hyperlinks
        *[{
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_idx,
                    "endRowIndex": row_idx + 1,
                    "startColumnIndex": col_idx,
                    "endColumnIndex": col_idx + 1
                },
                "cell": {
                    "userEnteredValue": {
                        "formulaValue": f'=HYPERLINK("{url[1]}"; "{url[0]}")'
                    },
                    "userEnteredFormat": {
                        "textFormat": {
                            "foregroundColor": {"red": 0.0, "green": 0.0, "blue": 1.0},
                            "underline": True
                        }
                    }
                },
                "fields": "userEnteredValue,userEnteredFormat.textFormat"
                }
        } for (row_idx, col_idx), url in hyperlink_map.items()],
            # 4) notes from `comments` dict
        *[{
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row_idx,
                    "endRowIndex": row_idx + 1,
                    "startColumnIndex": col_idx,
                    "endColumnIndex": col_idx + 1
                },
                "cell": {"note": str(note)},
                "fields": "note"
            }
        } for (row_idx, col_idx), note in comments.items()]
    ]
 
    sheets_service.spreadsheets().batchUpdate(
        spreadsheetId=MASTER_SPREADSHEET_ID,
        body={"requests": requests}
    ).execute()

def color_master_table_cell(row: int, col: int, status: str = None):

    sheets_service = GoogleApiService().sheets_service
    sheet_id = _get_master_table_sheet_id()
    white = {"red": 1.0, "green": 1.0, "blue": 1.0}

    body = {
    "requests": [
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row - 1,
                    "endRowIndex": row,
                    "startColumnIndex": col - 1,
                    "endColumnIndex": col
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": OperationManager.COLORS_DICT.get(status, white) if status else white
                    }
                },
                "fields": "userEnteredFormat.backgroundColor"
                }
            }
        ]
    }

    sheets_service.spreadsheets().batchUpdate(
        spreadsheetId=MASTER_SPREADSHEET_ID,
        body=body
    ).execute()

def convert_dttm(dttm: str) -> datetime:

    try:
        return datetime.strptime(dttm, '%Y-%m-%dT%H:%M')
    except ValueError:
        pass

    try:
        return datetime.strptime(dttm, '%Y-%m-%dT%H:%M:%S.%fZ')
    except ValueError:
        pass

def _make_project_rows_for_sheets(project_data: list[dict]) -> list[list]:

    result = []
    last_project_operation = None
    closed_job_cards = 0
    groups = ['COAT', 'MECH', 'OTK', 'NDT', 'CEX_TO', 'OUT', 'PPK']
    stats_dict =  {'in_work' : 0, 'closed' : 0, 'to_do' : 0}
    now = datetime.now()

    project_stats = {group : stats_dict.copy() for group in groups}

    for job_card in project_data:
        
        job_card_stats = {group : stats_dict.copy() for group in groups}
        operations = job_card.get('operations')
        all_dttms = [convert_dttm(v[key])
                            for v in operations.values()
                            for key in ('start_dttm', 'end_dttm')
                            if v[key]]


        last_operation = max(all_dttms) if all_dttms else None

        if not last_project_operation and last_operation:
            last_project_operation = last_operation
        elif last_operation and last_operation > last_project_operation:
            last_project_operation = last_operation

        for op in operations:

            try:
                op_name = op if op == 'PPK' else re.sub(r'^\d+_', '', op)
            except IndexError:
                continue

            if operations[op]['end_dttm']:

                job_card_stats[op_name]['closed'] += 1
                project_stats[op_name]['closed'] += 1

                if op_name == 'PPK':
                    closed_job_cards +=1

            elif operations[op]['start_dttm']:

                job_card_stats[op_name]['in_work'] += 1
                project_stats[op_name]['in_work'] += 1

            else:

                job_card_stats[op_name]['to_do'] += 1
                project_stats[op_name]['to_do'] += 1

        result.append([
            job_card.get('project', ''),
            job_card.get('name', ''),
            *['/'.join(map(str, values_dict.values())) for values_dict in job_card_stats.values()],
            round((now - last_operation).total_seconds() / 3600, 2) if last_operation else ''
        ])

    project_row = [[
        project_data[0].get('project'),
        f'{closed_job_cards}/{len(project_data)}',
        *['/'.join(map(str, values_dict.values())) for values_dict in project_stats.values()],
        round((now - last_project_operation).total_seconds() / 3600, 2) if last_project_operation else ''
    ]]

    return project_row

def update_projects_google_sheet():

    now_str = datetime.now(tz=ZoneInfo("Europe/Moscow")).strftime('%d.%m.%Y %H:%M')

    result = [
        ['Last update', now_str, 'COAT', 'MECH', 'OTK', 'NDT', 'CEX_TO', 'OUT', 'PPK', 'HOURS_FROM_LAST_OP']
    ]
    sheets_service = GoogleApiService().sheets_service
    job_card_data = OperationManager()._get_operations_data()

    sheet_metadata = sheets_service.spreadsheets().get(
        spreadsheetId=MASTER_SPREADSHEET_ID,
        fields="sheets(properties(sheetId,title))"
    ).execute()

    sheet_id =  next(
        sheet['properties']['sheetId'] 
        for sheet in sheet_metadata['sheets']
        if sheet['properties']['title'] == 'projects'
    )

    with DbManager() as db:
        projects, _ = db.execute_query('SELECT DISTINCT project FROM job_cards WHERE is_active = True ORDER BY project', fetch_results=True)

    
    for row in projects:
        project = row[0]
        result.extend(
            _make_project_rows_for_sheets(list(filter(lambda x: x['project'] == project, job_card_data)))
        )

    requests = [
        {
            "updateCells": {
                "range": {"sheetId": sheet_id},
                "fields": "userEnteredValue,userEnteredFormat,note"
                }
        },
        # 1. Update all cell values
        {
            "updateCells": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": len(result),
                    "startColumnIndex": 0,
                    "endColumnIndex": max([len(row) for row in result]) if result else 0
                },
                "rows": [
                        {
                            "values": [
                                {"userEnteredValue": {"stringValue": str(cell)}}
                                for cell in row
                            ]
                        }
                        for row in result
                    ],
                "fields": "userEnteredValue"
            }
        }
    ]

    sheets_service.spreadsheets().batchUpdate(
        spreadsheetId=MASTER_SPREADSHEET_ID,
        body={"requests": requests}
    ).execute()

def place_note_master_table(row: int, col: int, note:str = None):

    sheets_service = GoogleApiService().sheets_service

    body = {
        "requests": [{
            "repeatCell": {
                "range": {
                    "sheetId": _get_master_table_sheet_id(),
                    "startRowIndex": row,
                    "endRowIndex": row + 1,
                    "startColumnIndex": col,
                    "endColumnIndex": col + 1
                },
                "cell": {"note": note},        # -> null in JSON clears the note
                "fields": "note"
            }
        }]
    }
    return sheets_service.spreadsheets().batchUpdate(
        spreadsheetId=MASTER_SPREADSHEET_ID, body={"requests": body["requests"]}
    ).execute()

def close_job_card_color(row: int):

    sheets_service = GoogleApiService().sheets_service
    sheet_id = _get_master_table_sheet_id()

    body = {
    "requests": [
        {
            "repeatCell": {
                "range": {
                    "sheetId": sheet_id,
                    "startRowIndex": row - 1,
                    "endRowIndex": row,
                    "startColumnIndex": 0,
                    "endColumnIndex": 5
                },
                "cell": {
                    "userEnteredFormat": {
                        "backgroundColor": OperationManager.COLORS_DICT.get('completed')
                    }
                },
                "fields": "userEnteredFormat.backgroundColor"
                }
            }
        ]
    }

    sheets_service.spreadsheets().batchUpdate(
        spreadsheetId=MASTER_SPREADSHEET_ID,
        body=body
    ).execute()
