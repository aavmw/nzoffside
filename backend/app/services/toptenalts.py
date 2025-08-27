def top_ten_alts():
    import pandas as pd
    import re, gspread, os
    from openpyxl.styles import Border, Side, PatternFill

    # Определяем области доступа
    SCOPES = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]

    creds = {    }

    client = gspread.service_account_from_dict(creds)

    # Открытие таблицы с заказми
    table_id = "1wa0SRbz2G3OLp8IxRAJ-nBOTrd4cE83AJdxLQ9mddu4"
    sheet_id = "1903516609"
    sheet = client.open_by_key(table_id).get_worksheet_by_id(sheet_id)
    # Получение всех записей в DataFrame
    data = sheet.get_all_values()
    orders = pd.DataFrame(data[1:], columns=data[0])

    # Открытие таблицы с альтернативами
    table_id2 = "1eOfwP72KXxMaBKyGNTLXQw99-l8vDH5x4b81K9Hftgc"
    sheet_id2 = "482659015"
    sheet2 = client.open_by_key(table_id2).get_worksheet_by_id(sheet_id2)
    data2 = sheet2.get_all_values()
    all_alts = pd.DataFrame(data2[1:], columns=data2[0])

    # Открытие таблицы Aviator
    table_id3 = "1pl0szvv0iE5boAziGGpWyBi4inUUPRqntQaxRPuRV2Y"
    sheet_id3 = "0"
    sheet3 = client.open_by_key(table_id3).get_worksheet_by_id(sheet_id3)
    data3 = sheet3.get_all_values()
    aviator = pd.DataFrame(data3[1:], columns=data3[0])

    # Открытие таблицы со всеми партийными номерами
    table_id8 = "1eOfwP72KXxMaBKyGNTLXQw99-l8vDH5x4b81K9Hftgc"
    sheet_id8 = "2049102023"
    sheet8 = client.open_by_key(table_id8).get_worksheet_by_id(sheet_id8)
    data8 = sheet8.get_all_values()
    all_pn = pd.DataFrame(data8[1:], columns=data8[0])

    # Открытие таблицы с контактами менеджера
    table_id9 = "1Lkz9yC485IU7tBeJlQKX4y8kTq2Uk7etSDrbX_hdlwk"
    sheet_id9 = "1691816393"
    sheet9 = client.open_by_key(table_id9).get_worksheet_by_id(sheet_id9)
    data9 = sheet9.get_all_values()
    contacts = pd.DataFrame(data9[1:], columns=data9[0])

    contacts.columns.values[4] = "Телеграмм "

    # Убираем финализированные пары
    all_alts = all_alts[all_alts["final"] != "TRUE"]
    all_alts = all_alts[all_alts["final"] != "true"]

    # Создаём список альтов, которые сами являются НЗ
    NZ_list = all_pn[all_pn["marks"] == "NZ"]
    NZ_list = NZ_list["part_number"].to_list()

    all_alts = all_alts[["main_pn", "alt_pn", "alt_status"]]

    # Получаем список подтвержденных альтов
    confirmed_list = all_alts[all_alts["alt_status"] == "confirmed"]
    confirmed_list = confirmed_list.drop_duplicates(subset="alt_pn")
    confirmed_list = confirmed_list["alt_pn"].to_list()

    # Получаем фрейм только с опциональными, если альт в списке подтвержденных к другому парту, то удаляем
    optional_alts = all_alts[all_alts["alt_status"] == "optional"]
    optional_alts = optional_alts[~optional_alts["alt_pn"].isin(confirmed_list)]
    optional_alts = optional_alts[~optional_alts["alt_pn"].isin(NZ_list)]

    # Получаем список опциональных альтов
    optional_list = optional_alts.drop_duplicates(subset="alt_pn")
    optional_list = optional_list["alt_pn"].to_list()

    alt_orders = orders[orders["part_number"].isin(optional_list)]
    alt_orders = alt_orders[alt_orders["order_status"] != "cancelled"]
    alt_orders = alt_orders[alt_orders["order_status"] != "warranty"]

    # Оставляем только нужные столбцы
    alt_orders = alt_orders[["part_number", "customer"]]

    # Приводим к единому регистру
    alt_orders["customer"] = alt_orders["customer"].str.upper()

    # убираем закупки на склад
    alt_orders = alt_orders[alt_orders["customer"] != "STOCK"]
    alt_orders = alt_orders[
        ~alt_orders["customer"].str.contains(
            r"\bAC[\s\-]*DC\b", flags=re.IGNORECASE, na=False
        )
    ]

    alt_orders["customer"] = alt_orders["customer"].replace({"CRUISER PHIL": "CRUISER"})
    alt_orders["customer"] = alt_orders["customer"].replace({"APORT": "AESHAM"})
    alt_orders["customer"] = alt_orders["customer"].replace({"ROCK STAR": "ROVNU"})

    # Группируем по парт номеру, считаем количество покупок
    alt_orders_group = alt_orders.groupby("part_number", as_index=False)[
        "customer"
    ].agg("count")

    # Переименуем столбец с количеством
    alt_orders_group = alt_orders_group.rename(
        columns={"customer": "count_buys", "part_number": "alt_pn"}
    )

    # Оставляем уникальную пару альт номер - клиент
    alt_buyers = alt_orders.drop_duplicates().reset_index(drop=True)

    # Создаем список альтов, на которые есть заказы
    activ_alts_list = alt_orders_group["alt_pn"].to_list()

    # Оставляем только запросы с партами, на которые есть заказы
    aviator = aviator[aviator["part_number"].isin(activ_alts_list)].reset_index(
        drop=True
    )
    aviator["rfq_company"] = aviator["rfq_company"].str.upper()
    aviator = aviator[["part_number", "rfq_company"]]
    aviator = aviator.rename(columns={"rfq_company": "customer"})
    aviator = aviator.drop_duplicates().reset_index(drop=True)

    aviator["customer"] = aviator["customer"].replace({"BTC": "BTC-J"})
    aviator["customer"] = aviator["customer"].replace({"CRUISER PHIL": "CRUISER"})
    aviator["customer"] = aviator["customer"].replace({"CZ-T": "CZT"})
    aviator["customer"] = aviator["customer"].replace({"AESHAM2": "AESHAM"})
    aviator["customer"] = aviator["customer"].replace({"APORT": "AESHAM"})
    aviator["customer"] = aviator["customer"].replace({"ROCK STAR": "ROVNU"})

    aviator = aviator[
        ~aviator["customer"].str.contains(
            r"\bAC[\s\-]*DC\b", flags=re.IGNORECASE, na=False
        )
    ]

    # Получаем плный список всех, кто покупал или запрашивал альт
    alt_final = pd.concat([alt_buyers, aviator], ignore_index=True)
    alt_final = alt_final.drop_duplicates()
    alt_final = alt_final.rename(columns={"part_number": "alt_pn"})

    all_alts = all_alts[all_alts["alt_status"] != "confirmed"]

    # Получаем список основных партов для исследуемых партийников
    all_alts = all_alts[all_alts["alt_pn"].isin(activ_alts_list)]
    main_NZ_list = all_alts.drop_duplicates(subset="main_pn")
    # Получаем список основных партов для фильтрованных опциональных, по которым данные уже собраны
    main_NZ_list = all_alts["main_pn"].to_list()

    # Оставляем только заказы по партам НЗ изи списка
    nz_orders = orders[orders["part_number"].isin(main_NZ_list)]
    nz_orders = nz_orders[nz_orders["order_status"] != "cancelled"]
    nz_orders = nz_orders[nz_orders["order_status"] != "warranty"]

    # Оставляем только нужные столбцы
    nz_orders = nz_orders[["part_number", "customer"]]

    # Приводим к единому регистру
    nz_orders["customer"] = nz_orders["customer"].str.upper()

    # убираем закупки на склад
    nz_orders = nz_orders[nz_orders["customer"] != "STOCK"]
    nz_orders = nz_orders[
        ~nz_orders["customer"].str.contains(
            r"\bAC[\s\-]*DC\b", flags=re.IGNORECASE, na=False
        )
    ]
    nz_orders = nz_orders.drop_duplicates().reset_index(drop=True)
    nz_orders = nz_orders.rename(columns={"part_number": "main_pn"})

    # Оставляем только уникальные пары парт-альт
    all_alts = all_alts.drop(columns="alt_status").reset_index(drop=True)

    nz_orders = nz_orders.merge(all_alts, on="main_pn", how="left")

    alt_final = alt_final.merge(all_alts, on="alt_pn", how="left")

    final_df = pd.concat([nz_orders, alt_final])

    final_df = final_df.dropna()
    final_df = final_df.drop_duplicates().reset_index(drop=True)

    final_df = final_df.merge(alt_orders_group, on="alt_pn", how="left")

    final_df = final_df[
        ~final_df["customer"].str.contains(
            r"\bAC[\s\-]*DC\b", flags=re.IGNORECASE, na=False
        )
    ]

    final_df = final_df[["main_pn", "alt_pn", "customer", "count_buys"]]

    final_df = final_df[final_df["main_pn"].isin(NZ_list)]

    final_df = final_df.rename(
        columns={
            "main_pn": "Основной PN",
            "alt_pn": "Альтернативный PN",
            "count_buys": "Всего покупок альта",
        }
    )

    # Приводим контакты менеджеров к общему виду
    contacts = contacts.rename(columns={"Никнейм клиента": "Link"})
    contacts["Link"] = (
        contacts["Link"].astype(str).replace({" ": "_", "-": "_"}, regex=True)
    )

    contacts = contacts.drop_duplicates(subset="Link").reset_index(drop=True)

    contacts_list = contacts["Link"].to_list()

    # Приводим все значения в 'customer' к строковому типу и заменяем пробелы и дефисы
    final_df["customer"] = (
        final_df["customer"].astype(str).replace({"  ": " "}, regex=True)
    )
    final_df["customer"] = (
        final_df["customer"].astype(str).replace({" ": "_", "-": "_"}, regex=True)
    )
    final_df["customer"] = final_df["customer"].replace({"CRUISER_PHIL": "CRUISER"})
    final_df["customer"] = final_df["customer"].replace({"BTC": "BTC_J"})
    final_df["customer"] = final_df["customer"].replace({"APORT": "AESHAM"})
    final_df["customer"] = final_df["customer"].replace({"ROCK STAR": "ROVNU"})
    # Удаляем лишние пробелы по краям в столбце 'customer'
    final_df["customer"] = final_df["customer"].str.strip()
    final_df = final_df[final_df["customer"].isin(contacts_list)]

    # Удаляем дубликаты по столбцу 'customer'
    clients_list = final_df.drop_duplicates(subset="customer").reset_index(drop=True)

    # Преобразуем столбец в список
    clients_list = clients_list["customer"].to_list()

    clients_data = {}

    for client in clients_list:
        client_data = final_df[final_df["customer"] == client]
        client_data = client_data.sort_values(by="Всего покупок альта", ascending=False)
        client_data = client_data.drop(columns="customer")
        clients_data[client] = client_data.head(10)

    # Преобразуем `contacts` в словарь для быстрого доступа
    contacts_dict = contacts.set_index("Link").to_dict(orient="index")

    file_name = "clients_top_10_alts.xlsx"
    path = os.path.dirname(os.path.abspath(__file__)) + "/" + file_name

    # Список для хранения ссылок на каждый лист и информации о клиентах
    index_data_list = []

    # Используем ExcelWriter для записи нескольких DataFrame в разные листы
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Перебираем клиентов и создаём страницы
        for link, data in clients_data.items():
            # Записываем DataFrame в Excel на лист клиента
            data.to_excel(writer, sheet_name=link, index=False)

            # Получаем информацию о клиенте
            client_info = contacts_dict.get(link, {})
            director = client_info.get("Руководитель", "")  # Руководитель
            telegram = client_info.get("Телеграмм", "")  # Телеграмм
            manager = client_info.get(
                "Менеджер ответственный за клиента ", ""
            )  # Менеджер
            telegram2 = client_info.get("Телеграмм ", "")  # Телеграмм
            # Добавляем запись с ссылкой на лист и информацию о клиенте для главной страницы
            index_data_list.append(
                {
                    "Link": f'=HYPERLINK("#{link}!A1", "{link}")',
                    "Руководитель": director,
                    "Телеграмм": telegram,
                    "Менеджер ответственный за клиента": manager,
                    "Телеграмм ": telegram2,
                }
            )

            # Получаем объект книги и листа для последующей настройки
            worksheet = writer.sheets[link]

            # Устанавливаем ширину столбцов в 20
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter  # Получаем буквенное имя столбца
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = max_length + 2
                worksheet.column_dimensions[column].width = max(adjusted_width, 20)

            # Увеличиваем ширину столбца E в 6 раз на всех листах, кроме главного
            if link != "Список_компаний":
                worksheet.column_dimensions["E"].width = (
                    worksheet.column_dimensions["E"].width * 4
                )

            # Добавим информацию о клиенте в последнюю строку на странице клиента
            client_info_start_col = (
                len(data.columns) + 2
            )  # Начальный столбец для данных клиента
            worksheet.cell(
                row=1, column=client_info_start_col, value="Информация о клиенте"
            )  # Заголовок
            header_fill = PatternFill(
                start_color="D3D3D3", end_color="D3D3D3", fill_type="solid"
            )
            worksheet.cell(row=1, column=client_info_start_col).fill = header_fill

            # Заполняем детали клиента построчно
            for idx, (key, value) in enumerate(client_info.items(), start=2):
                worksheet.cell(
                    row=idx, column=client_info_start_col, value=f"{key}: {value}"
                )

            # Настроим границы ячеек (только для первых трех столбцов на каждом листе)
            thin_border = Border(
                left=Side(border_style="thin", color="000000"),
                right=Side(border_style="thin", color="000000"),
                top=Side(border_style="thin", color="000000"),
                bottom=Side(border_style="thin", color="000000"),
            )
            for row in worksheet.iter_rows(
                min_col=1, max_col=3
            ):  # Ограничиваем только первыми тремя столбцами
                for cell in row:
                    cell.border = thin_border

            # Добавляем серую заливку для заголовков
            if link != "Список_компаний":  # Для всех листов, кроме главного
                for cell in worksheet[1][
                    :3
                ]:  # Первая строка с заголовками, только для первых трех столбцов
                    cell.fill = header_fill
            else:  # Для главного листа (Список_компаний)
                for cell in worksheet[1]:  # Заголовки на первой строке
                    cell.fill = header_fill

            # Добавляем ссылку на "Список компаний" на каждом листе
            last_row = (
                len(data) + 2
            )  # +2 для учета заголовков и последней строки данных
            worksheet.cell(row=last_row, column=1, value="К списку компаний")
            worksheet.cell(row=last_row, column=1).hyperlink = "#Список_компаний!A1"

            # Окрашиваем ячейку с ссылкой "К списку компаний" в красный
            red_fill = PatternFill(
                start_color="ea9999", end_color="ea9999", fill_type="solid"
            )
            worksheet.cell(row=last_row, column=1).fill = red_fill

        # **Сортируем список по алфавиту перед созданием DataFrame**
        index_data_list.sort(
            key=lambda x: x["Link"].split('"')[1].lower()
        )  # Сортировка по названию листа

        # Создаём DataFrame из отсортированного списка для оглавления
        index_data = pd.DataFrame(index_data_list)

        # Записываем индексный лист
        index_data.to_excel(writer, sheet_name="Список_компаний", index=False)

        # Получаем доступ к листу с оглавлением
        index_sheet = writer.sheets["Список_компаний"]

        # Устанавливаем ширину столбцов для листа оглавления
        index_sheet.column_dimensions["A"].width = 40
        index_sheet.column_dimensions["B"].width = 30  # Для руководителя
        index_sheet.column_dimensions["C"].width = 30  # Для телеграма
        index_sheet.column_dimensions["D"].width = 40  # Для менеджера
        index_sheet.column_dimensions["E"].width = 30  # Для телеги менеджера

        # Настроим границы для оглавления (здесь оставляем все столбцы с границами)
        for row in index_sheet.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Добавляем серую заливку для заголовков в оглавлении
        for cell in index_sheet[1]:  # Заголовки на первой строке
            cell.fill = header_fill

        # Перемещаем лист с оглавлением в начало (первым листом)
        writer.book._sheets.sort(key=lambda sheet: sheet.title != "Список_компаний")

    return path
